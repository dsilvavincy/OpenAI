"""
Database T12 Workbook Processor

Handles the "Portfolio Database" format where:
1. Financials are in a sheet named "[Property]-Fin"
2. Budgets are in a sheet named "[Property]-Bgt"
3. Data is structured as a time-series database (Metrics in Col 1, Dates in Col 2+)
"""
import pandas as pd
import numpy as np
import openpyxl
import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path
from .base_processor import BaseFormatProcessor

class DatabaseT12Processor(BaseFormatProcessor):
    """
    Processor for CRES Portfolio Database format.
    """
    
    def __init__(self):
        super().__init__(
            format_name="Database_T12_Workbook",
            format_description="Dynamic database format with separated -Fin and -Bgt sheets"
        )
    
    def can_process(self, file_path: Path, sheet_name: Optional[str] = None) -> bool:
        """
        Check if file follows the Database T12 format (contains *-Fin and *-Bgt sheets).
        """
        try:
            # If file-like object, reset pointer
            if hasattr(file_path, 'seek'):
                file_path.seek(0)
            wb = openpyxl.load_workbook(file_path, read_only=True, keep_links=False)
            sheet_names = wb.sheetnames
            wb.close()
            
            # Check for at least one pair of matching -Fin and -Bgt sheets
            fin_sheets = [s for s in sheet_names if s.endswith("-Fin")]
            if not fin_sheets:
                return False
                
            for fin in fin_sheets:
                base_name = fin[:-4] # Remove "-Fin"
                if f"{base_name}-Bgt" in sheet_names:
                    return True
            
            return False
            
        except Exception:
            return False

    def process(self, file_path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Process the workbook and return a unified DataFrame with Actuals, Budgets, and YTD.
        """
        # If file-like object, reset pointer
        if hasattr(file_path, 'seek'):
            file_path.seek(0)
            
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        all_sheets = wb.sheetnames
        
        # Read property whitelist from DB sheet
        property_whitelist = self._read_property_whitelist(wb)
        
        # Identify pairs (handle both "Name-Fin" and "Name -Fin" patterns)
        pairs = []
        for s in all_sheets:
            if s.endswith("-Fin") or s.endswith(" -Fin"):
                # Remove suffix to get base name
                if s.endswith(" -Fin"):
                    base_name = s[:-5]  # Remove " -Fin"
                    bgt_sheet = f"{base_name} -Bgt"
                else:
                    base_name = s[:-4]  # Remove "-Fin"
                    bgt_sheet = f"{base_name}-Bgt"
                
                if bgt_sheet in all_sheets:
                    pairs.append((base_name, s, bgt_sheet))
        
        # Filter by whitelist (case-insensitive match)
        if property_whitelist:
            # Create lowercase lookup for case-insensitive matching
            whitelist_lower = {prop.lower(): prop for prop in property_whitelist}
            pairs = [(base, fin, bgt) for base, fin, bgt in pairs 
                     if base.lower() in whitelist_lower]
        
        if not pairs:
            wb.close()
            raise ValueError("No valid Property-Fin/Property-Bgt sheet pairs found.")
            
        final_frames = []
        
        for property_name, fin_sheet, bgt_sheet in pairs:
            # --- Process Financials (Actuals) ---
            if hasattr(file_path, 'seek'):
                file_path.seek(0)
            df_fin = pd.read_excel(file_path, sheet_name=fin_sheet, header=None, engine='openpyxl')
            if len(df_fin) < 8:
                continue
                
            # Header is Row 7 (Index 6)
            header_row = df_fin.iloc[6]
            # Verify Header Content (Col 0 should imply Metric)
            if not (isinstance(header_row[0], str) and ("Actuals" in header_row[0] or "Metric" in header_row[0])):
                # Try to find header if not exactly at 7? User said row 7. Assume row 7.
                pass

            data_fin = df_fin.iloc[7:].copy()
            
            # Identify Date Columns (Col 1 onwards)
            # Filter columns that parse to datetime
            date_col_map = {} # {col_idx: datetime}
            for idx, val in enumerate(header_row):
                if idx == 0: continue # Metric column
                dt = self._parse_header_date(val)
                if dt:
                    date_col_map[idx] = dt
            
            if not date_col_map:
                continue
                
            # Rename columns: Col 0 -> Metric, others -> Date
            cols = list(data_fin.columns)
            rename_dict = {cols[0]: "Metric"}
            for idx, dt in date_col_map.items():
                rename_dict[cols[idx]] = dt
            
            # Determine meaningful cutoff date for Actuals based on "Net Eff. Gross Income"
            # This is the most reliable indicator of whether a month has actuals.
            
            # Find the row index for Net Eff. Gross Income
            negi_idx = data_fin[data_fin[cols[0]].astype(str).str.contains("Net Eff. Gross Income", case=False, na=False)].index
            
            valid_date_cols = []
            
            if not negi_idx.empty:
                target_idx = negi_idx[0]
                # Check only this row for each column
                for col_idx, dt in date_col_map.items():
                    col_name = cols[col_idx]
                    val = pd.to_numeric(data_fin.loc[target_idx, col_name], errors='coerce')
                    if pd.notna(val) and val != 0:
                        valid_date_cols.append((col_idx, dt))
            else:
                # Fallback: Check "Total Income" or strictly >5 non-zeros if NEGI missing
                for col_idx, dt in date_col_map.items():
                    col_name = cols[col_idx]
                    col_data = pd.to_numeric(data_fin[col_name], errors='coerce').fillna(0)
                    if (col_data != 0).sum() > 5: 
                        valid_date_cols.append((col_idx, dt))
            
            # Sort by date
            valid_date_cols.sort(key=lambda x: x[1])
            
            # GUARDRAIL: Remove dates in the future (Actuals cannot be in future)
            # We allow the current month (even if day > today), but not future months.
            current_date = datetime.datetime.now()
            
            def is_past_or_current_month(dt):
                if dt.year < current_date.year:
                    return True
                if dt.year == current_date.year and dt.month <= current_date.month:
                    return True
                return False
                
            valid_date_cols = [x for x in valid_date_cols if is_past_or_current_month(x[1])]
            
            # If we found valid columns, take the set up to the last valid one
            # Actually, standard practice: If Nov-25 is empty, but Jan-25 is full, we keep Jan-25.
            # We should keep columns up to the MAX date that has meaningful data.
            
            max_valid_date = None
            if valid_date_cols:
                max_valid_date = valid_date_cols[-1][1]
            
            # Calculate T12 window (last 12 months from max_valid_date)
            # This reduces payload size by excluding older history
            t12_start_date = None
            if max_valid_date:
                from dateutil.relativedelta import relativedelta
                t12_start_date = max_valid_date - relativedelta(months=11)  # 11 months back + current = 12 total
            
            # Re-build date_col_map to only include dates <= max_valid_date AND >= t12_start_date
            final_date_col_map = {}
            if max_valid_date:
                for idx, dt in date_col_map.items():
                    if dt <= max_valid_date:
                        if t12_start_date is None or dt >= t12_start_date:
                            final_date_col_map[idx] = dt
            else:
                final_date_col_map = date_col_map # Fallback
            
            keep_cols = [cols[i] for i in [0] + list(final_date_col_map.keys())]
            data_fin = data_fin[keep_cols].rename(columns={cols[0]: "Metric", **{cols[k]: v for k,v in final_date_col_map.items()}})
            
            # Drop rows with empty Metric
            data_fin = data_fin.dropna(subset=["Metric"])
            
            # Formatting: Clean up metric names (e.g. 0.05 -> "Valuation p/unit 5%")
            data_fin["Metric"] = data_fin["Metric"].apply(self._format_metric_name)
            
            # --- Differentiate Duplicate Metrics (e.g., Financial Data Section) ---
            # There is a distinct "Financial Data" section at the bottom which re-uses metric names.
            # We must detect this section and rename metrics to avoid collisions (e.g., Gross Scheduled Rent).
            fd_idx = data_fin[data_fin['Metric'].astype(str).str.contains('Financial Data', case=False, na=False)].index
            if not fd_idx.empty:
                cutoff_stats = fd_idx[0]
                # Identify rows AFTER the Financial Data header
                # Since dataframe index preserves original row numbers, we can use simple comparison
                stats_mask = data_fin.index > cutoff_stats
                # Append suffix
                data_fin.loc[stats_mask, "Metric"] = data_fin.loc[stats_mask, "Metric"] + " (Stats)"
            
            # Melt Actuals
            actual_long = data_fin.melt(id_vars="Metric", var_name="Period", value_name="Value")
            actual_long["Value"] = pd.to_numeric(actual_long["Value"], errors='coerce').fillna(0)
            
            # Identify valid metrics for YTD calculation (stop at "Monthly Cash Flow")
            # We do this on data_fin before melting to preserve order
            valid_ytd_metrics = set()
            m_cf_idx_fin = data_fin[data_fin['Metric'].astype(str).str.contains('Monthly Cash Flow', case=False, na=False)].index
            if not m_cf_idx_fin.empty:
                cutoff_idx = m_cf_idx_fin[0]
                # data_fin is indexed by row numbers from read_excel, let's just take head
                # Since we dropped na metrics, index might be non-contiguous, so use .loc
                # Actually, simpler: just iterate or finding position
                # Get all metrics in order
                all_metrics_ordered = data_fin['Metric'].tolist()
                try:
                    # Find index in list (first occurrence)
                    # Note: formatted names might differ slightly, but we just formatted them.
                    # Searching for substring match in list might be safer?
                    # But we used the index directly above.
                    # Wait, data_fin index is preserved. cutoff_idx is the index label.
                    # So we can effectively toggle based on index labels assuming increasing
                    valid_mask = data_fin.index <= cutoff_idx
                    valid_ytd_metrics = set(data_fin[valid_mask]['Metric'].unique())
                except:
                    # Fallback: keep all
                    valid_ytd_metrics = set(all_metrics_ordered)
            else:
                valid_ytd_metrics = set(data_fin['Metric'].unique())

            # --- Process Budgets ---
            if hasattr(file_path, 'seek'):
                file_path.seek(0)
            df_bgt = pd.read_excel(file_path, sheet_name=bgt_sheet, header=None, engine='openpyxl')
            # Assuming aligned structure, but safe to parse dates again
            header_row_bgt = df_bgt.iloc[6]
            data_bgt = df_bgt.iloc[7:].copy()
            
            date_col_map_bgt = {}
            for idx, val in enumerate(header_row_bgt):
                if idx == 0: continue
                dt = self._parse_header_date(val)
                if dt:
                    date_col_map_bgt[idx] = dt
            
            # Filter Budget columns to match Actuals cutoff (prevent future budgets AND respect T12 window)
            final_date_col_map_bgt = {}
            if max_valid_date:
                for idx, dt in date_col_map_bgt.items():
                    if dt <= max_valid_date:
                        if t12_start_date is None or dt >= t12_start_date:
                            final_date_col_map_bgt[idx] = dt
            else:
                final_date_col_map_bgt = date_col_map_bgt
            
            cols_bgt = list(data_bgt.columns)
            rename_dict_bgt = {cols_bgt[0]: "Metric"}
            for idx, dt in final_date_col_map_bgt.items():
                rename_dict_bgt[cols_bgt[idx]] = dt
                
            keep_cols_bgt = [cols_bgt[i] for i in [0] + list(final_date_col_map_bgt.keys())]
            data_bgt = data_bgt[keep_cols_bgt].rename(columns=rename_dict_bgt)
            
            # Drop rows with empty Metric
            data_bgt = data_bgt.dropna(subset=["Metric"])

            # Formatting: Clean up metric names
            data_bgt["Metric"] = data_bgt["Metric"].apply(self._format_metric_name)
            
            # Apply Budget Nullification Logic (Monthly Cash Flow cutoff)
            # Find "Monthly Cash Flow"
            m_cf_series = data_bgt['Metric'].astype(str).str.lower()
            m_cf_found = m_cf_series.str.contains('monthly cash flow', na=False)
            
            if m_cf_found.any():
                cutoff_index = m_cf_found.idxmax() # Get index of first True
                # Nullify all values for rows numerically > cutoff_index
                rows_to_null = data_bgt.index[data_bgt.index > cutoff_index]
                if not rows_to_null.empty:
                    # Set date columns to NaN
                    date_cols_bgt_list = [c for c in data_bgt.columns if c != "Metric"]
                    data_bgt.loc[rows_to_null, date_cols_bgt_list] = np.nan

            # Melt Budgets
            budget_long = data_bgt.melt(id_vars="Metric", var_name="Period", value_name="BudgetValue")
            budget_long["BudgetValue"] = pd.to_numeric(budget_long["BudgetValue"], errors='coerce').fillna(0)
            
            # --- Merge ---
            # Merge on Metric and Period
            combined = pd.merge(actual_long, budget_long, on=["Metric", "Period"], how="outer")
            
            # Ensure no NaNs are introduced by the outer merge
            combined["Value"] = combined["Value"].fillna(0)
            combined["BudgetValue"] = combined["BudgetValue"].fillna(0)
            
            # GUARDRAIL: Brute-force sweep for extreme overflow artifacts (e.g. INT64_MIN used for NaNs in some engines)
            # This squashes the -9.22E+18 values reported by the user
            for col in ["Value", "BudgetValue"]:
                mask_extreme = (combined[col] < -1e15) | (combined[col] > 1e15)
                if mask_extreme.any():
                    combined.loc[mask_extreme, col] = 0
            
            combined["Property"] = property_name
            combined["Sheet"] = fin_sheet # Metadata
            combined["IsYTD"] = False
            
            # Ensure Period is datetime for .dt accessor
            combined["Period"] = pd.to_datetime(combined["Period"], errors='coerce')
            
            combined["MonthParsed"] = combined["Period"]
            combined["Month"] = combined["Period"].dt.month
            combined["Year"] = combined["Period"].dt.year
            combined["Month_Name"] = combined["Period"].dt.month_name()
            
            # Filter out rows where BOTH Actual and Budget are zero/nan
            # This cleans up the dataset significantly
            combined = combined[~((combined["Value"] == 0) & (combined["BudgetValue"].fillna(0) == 0))]
            
            # --- Calculate YTD ---
            # We must calculate YTD for the LATEST available month (or all months?)
            # Standard processor generally provides YTD as a separate "Period='YTD'" row.
            
            if not combined.empty:
                # Find latest date in the data
                valid_dates = [d for d in combined['Period'].unique() if isinstance(d, datetime.datetime)]
                if valid_dates:
                    latest_date = max(valid_dates)
                    current_year = latest_date.year
                    
                    # Filter for current year
                    cy_data = combined[
                        combined['Period'].apply(lambda x: isinstance(x, datetime.datetime) and x.year == current_year)
                    ]
                    
                    # Group by Metric and sum, ensure numeric only
                    # Warning: grouping on Metric can be tricky if names are not unique or messy.
                    # We assume names are standard.
                    ytd_values = cy_data.groupby("Metric")[["Value", "BudgetValue"]].sum().reset_index()
                    
                    # Filter YTD values to only include applicable Valid Metrics (above Monthly Cash Flow)
                    ytd_values = ytd_values[ytd_values['Metric'].isin(valid_ytd_metrics)]
                    
                    ytd_values["Period"] = "YTD"
                    ytd_values["Property"] = property_name
                    ytd_values["Sheet"] = fin_sheet
                    ytd_values["Sheet"] = fin_sheet
                    ytd_values["IsYTD"] = True
                    ytd_values["MonthParsed"] = pd.NaT
                    ytd_values["Month"] = np.nan
                    ytd_values["Year"] = np.nan
                    ytd_values["Month_Name"] = None
                    
                    # Combine Monthly + YTD
                    period_df = pd.concat([combined, ytd_values], ignore_index=True)
                    final_frames.append(period_df)
                else:
                    final_frames.append(combined)
            else:
                 final_frames.append(combined)

        wb.close()
        
        if not final_frames:
            return pd.DataFrame()
            
        return pd.concat(final_frames, ignore_index=True)

    def validate_format(self, df: pd.DataFrame) -> bool:
        """Validate processed DataFrame"""
        required = ["Metric", "Period", "Value", "BudgetValue", "Property"]
        if not all(col in df.columns for col in required):
            raise ValueError(f"Missing required columns. Found: {list(df.columns)}")
        return True

    def get_expected_metrics(self) -> List[str]:
        """Return expected metrics"""
        return ["Net Eff. Gross Income", "Total Expense", "EBITDA (NOI)"]

    def _format_metric_name(self, val: Any) -> str:
        """
        Format metric names, specifically converting decimal percentages 
        (e.g., 0.05) to readable strings (Valuation p/unit 5%).
        """
        if isinstance(val, (int, float)):
            # Check for common percentage-like values found in T12s
            # These often appear as row headers like 0.05, 0.06 etc.
            if 0 < val < 1:
                return f"Valuation p/unit {int(val*100)}%"
            return str(val)
        return str(val).strip()

    def _read_property_whitelist(self, wb) -> List[str]:
        """
        Read property names from 'DB' sheet, Column K, starting Row 2.
        Returns empty list if DB sheet doesn't exist or column is empty.
        """
        if 'DB' not in wb.sheetnames:
            return []  # No whitelist = process all
        
        try:
            db_sheet = wb['DB']
            properties = []
            
            # Column K is column 11 (1-indexed)
            row_idx = 2
            while row_idx < 1000:  # Safety limit
                cell_val = db_sheet.cell(row=row_idx, column=11).value
                if not cell_val or (isinstance(cell_val, str) and not cell_val.strip()):
                    break
                properties.append(str(cell_val).strip())
                row_idx += 1
            
            return properties
        except Exception:
            return []  # If any error, process all
    
    def _parse_header_date(self, val: Any) -> Optional[datetime.datetime]:
        """Attempt to parse openpyxl cell value as datetime"""
        if isinstance(val, datetime.datetime):
            return val
        if isinstance(val, str):
            try:
                # Try common formats
                return pd.to_datetime(val).to_pydatetime()
            except:
                pass
        return None
