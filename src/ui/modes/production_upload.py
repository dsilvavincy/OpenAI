"""
Production Upload - Streamlined file upload for production mode

Handles file upload with immediate processing and clean user feedback.
Focuses on the essential workflow without debug features.
"""
import streamlit as st
import pandas as pd
import os
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any


class ProductionUpload:
    """Production mode file upload handler."""
    
    def render(self, uploaded_file: Optional[Any], config: Dict[str, Any]):
        """
        Render the file upload section - streamlined for production.
        
        Args:
            uploaded_file: Current uploaded file object
            config: Configuration from sidebar
        """
        st.markdown("### 📁 Upload T12 Data")
        
        # Persistence Logic
        CACHE_DIR = Path(".cache")
        CACHE_DIR.mkdir(exist_ok=True)
        CACHE_MONTHLY = CACHE_DIR / "monthly_df.pkl"
        CACHE_YTD = CACHE_DIR / "ytd_df.pkl"
        
        # (Restore button removed as per user request)
        
        # Use session state to persist uploaded file
        if 'current_uploaded_file' not in st.session_state:
            st.session_state['current_uploaded_file'] = None
        
        uploaded_file = st.file_uploader(
            "Choose your T12 Excel file",
            type=['xlsx', 'xls', 'xlsm'],
            help="Upload your T12 property financial data for AI analysis",
            key="production_file_uploader"
        )
        
        # Only process if not already processed
        if uploaded_file is not None:
            st.session_state['current_uploaded_file'] = uploaded_file
            if 'processed_monthly_df' not in st.session_state or 'processed_ytd_df' not in st.session_state:
                # Progress bar instead of spinner
                progress_bar = st.progress(0, text="📂 Reading file...")
                
                progress_bar.progress(10, text="📊 Detecting format...")
                monthly_df, ytd_df, processed_data = self._handle_file_upload(uploaded_file, progress_bar)
                
                if monthly_df is not None:
                    progress_bar.progress(90, text="💾 Saving to cache...")
                    st.session_state['processed_monthly_df'] = monthly_df
                    st.session_state['processed_ytd_df'] = ytd_df
                    st.session_state['uploaded_file'] = uploaded_file
                    
                    if processed_data:
                        st.session_state['processed_data'] = processed_data
                    
                    # Cache to disk for persistence across refreshes
                    monthly_df.to_pickle(CACHE_MONTHLY)
                    ytd_df.to_pickle(CACHE_YTD)
                    
                    # Save to exports folder as requested
                    self._save_processed_data(monthly_df, ytd_df, uploaded_file.name)
                    
                    progress_bar.progress(100, text="✅ Complete!")
                    st.success("✅ File processed successfully! Monthly and YTD data available.")
                    st.rerun()  # Refresh to show results layout
                else:
                    progress_bar.empty()  # Clear progress bar on error
    
    def _handle_file_upload(self, uploaded_file, progress_bar=None):
        """
        Process uploaded T12 file using FormatRegistry for automatic detection.
        
        Args:
            uploaded_file: Streamlit uploaded file object
            progress_bar: Optional Streamlit progress bar to update
            
        Returns:
            Tuple of (monthly_df, ytd_df) or (None, None) if processing failed
        """
        try:
            from src.core.format_registry import FormatRegistry
            from src.utils.format_detection import store_detected_format
            import io
            
            # Reset buffer
            uploaded_file.seek(0)
            excel_buffer = io.BytesIO(uploaded_file.getvalue())
            
            if progress_bar:
                progress_bar.progress(20, text="🔍 Analyzing file structure...")
            
            # Use registry to detect and process
            registry = FormatRegistry()
            unified_df, processor = registry.process_file(excel_buffer)
            
            if progress_bar:
                progress_bar.progress(40, text="📊 Extracting data...")
            
            if unified_df is None or unified_df.empty:
                st.error("❌ No data could be extracted from the file.")
                return None, None, None
                
            # Store detected format for prompt selection
            store_detected_format(processor.format_name)
            
            if progress_bar:
                progress_bar.progress(60, text="🔄 Splitting Monthly/YTD...")
            
            # Split into monthly and YTD for the analysis engine
            monthly_df, ytd_df = self._split_unified_df(unified_df)
            
            # --- PORTFOLIO SNAPSHOT GENERATION ---
            if progress_bar:
                progress_bar.progress(70, text="📋 Generating Portfolio Snapshot...")
            processed_data = {}
            try:
                import openpyxl
                from src.core.report_generator import ReportGenerator
                
                # Re-open for specific cell extraction
                excel_buffer.seek(0)
                wb = openpyxl.load_workbook(excel_buffer, data_only=True)
                
                report_gen = ReportGenerator()
                
                # Iterate detected properties
                detected_props = monthly_df['Property'].unique()
                for prop in detected_props:
                     portfolio_html = report_gen.generate_portfolio_table(wb, prop)
                     processed_data[prop] = {
                         "portfolio_snapshot_html": portfolio_html
                     }
                     
                     # Extract RAW DataFrame for Export
                     if "CRES - Portfolio (Internal)" in wb.sheetnames:
                         ws_int = wb["CRES - Portfolio (Internal)"]
                         target_row = None
                         for row in ws_int.iter_rows(min_row=5, max_col=2):
                             cell_val = row[1].value
                             if cell_val and str(cell_val).strip().lower() == prop.strip().lower():
                                 target_row = row[0].row
                                 break
                         
                         if target_row:
                             headers = []
                             row_vals = []
                             for col in range(2, 31):
                                 h_val = ws_int.cell(row=4, column=col).value
                                 headers.append(str(h_val).strip() if h_val else f"Col_{col}")
                                 val = ws_int.cell(row=target_row, column=col).value
                                 
                                 # Basic formatting for export visibility
                                 if val is None: val = "-"
                                 elif isinstance(val, (int, float)):
                                     # Heuristic formatting
                                     h_str = headers[-1]
                                     if "Occupancy" in h_str or "Yield" in h_str or "vs" in h_str or "Seq" in h_str:
                                          if abs(val) < 5: val = f"{val:.1%}"
                                          else: val = f"{val:.1f}%"
                                     elif "DSCR" in h_str:
                                          val = f"{val:.2f}"
                                     else:
                                          val = f"{val:,.0f}" if val > 5 else str(val)
                                 
                                 row_vals.append(str(val))
                             
                             # Create DF
                             p_df = pd.DataFrame([row_vals], columns=headers)
                             processed_data[prop]["portfolio_snapshot_df"] = p_df
            except Exception as e:
                print(f"Portfolio Snapshot Error: {e}")
            # -------------------------------------
            
            if monthly_df is not None and not monthly_df.empty:
                metrics_count = monthly_df['Metric'].nunique()
                properties_count = monthly_df['Property'].nunique()
                st.success(f"✅ Processed {properties_count} properties, {metrics_count} unique metrics")
            else:
                st.error("❌ No monthly data found in file")
            
            if ytd_df is None or ytd_df.empty:
                st.warning("⚠️ No YTD data found in file")
                
            return monthly_df, ytd_df, processed_data
            
        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
            st.info("💡 Please ensure your file is a valid T12 Excel format")
            import traceback
            st.expander("Error Details").code(traceback.format_exc())
            return None, None, None

    def _split_unified_df(self, df):
        """Split unified DataFrame into monthly and YTD DataFrames for analysis."""
        if df is None or df.empty:
            return None, None
            
        # Split by the IsYTD flag
        monthly_df = df[df['IsYTD'] == False].copy()
        ytd_df = df[df['IsYTD'] == True].copy()
        
        # Fill in date info for YTD rows based on last month of each property
        # This aligns with how PropertyAnalyzer expects YTD data
        for prop in ytd_df['Property'].unique():
            prop_rows = monthly_df[monthly_df['Property'] == prop]
            if not prop_rows.empty:
                # Find the latest month for this property
                max_month_idx = prop_rows['MonthParsed'].idxmax()
                last_date_info = prop_rows.loc[max_month_idx]
                
                idx = ytd_df['Property'] == prop
                ytd_df.loc[idx, 'Month'] = last_date_info['Month']
                ytd_df.loc[idx, 'MonthParsed'] = last_date_info['MonthParsed']
                ytd_df.loc[idx, 'Year'] = last_date_info['Year']
                ytd_df.loc[idx, 'Month_Name'] = last_date_info['Month_Name']
        
        # Drop the IsYTD helper column if it exists to keep DataFrames clean
        if 'IsYTD' in monthly_df.columns:
            monthly_df = monthly_df.drop(columns=['IsYTD'])
        if 'IsYTD' in ytd_df.columns:
            ytd_df = ytd_df.drop(columns=['IsYTD'])
            
        return monthly_df, ytd_df

    def _save_processed_data(self, monthly_df, ytd_df, filename):
        """Save processed DataFrames to the exports folder."""
        import os
        from pathlib import Path
        from datetime import datetime
        
        try:
            export_dir = Path("exports")
            export_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = Path(filename).stem
            
            monthly_path = export_dir / f"{base_name}_monthly_{timestamp}.csv"
            ytd_path = export_dir / f"{base_name}_ytd_{timestamp}.csv"
            
            monthly_df.to_csv(monthly_path, index=False)
            ytd_df.to_csv(ytd_path, index=False)
            
            # Save a unified version for convenience
            unified_path = export_dir / f"{base_name}_unified_{timestamp}.csv"
            pd.concat([monthly_df, ytd_df], ignore_index=True).to_csv(unified_path, index=False)
            
            st.info(f"📁 Processed data exported to `{export_dir}`")
            
        except Exception as e:
            st.warning(f"⚠️ Could not save CSVs to exports folder: {str(e)}")
    
