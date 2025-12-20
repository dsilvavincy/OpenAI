"""Debug script to inspect whitelist and property matching."""
import sys
from pathlib import Path
import openpyxl
from rich.console import Console
from rich.table import Table

project_root = Path(__file__).parent.parent
sys.path.append(str(project_root))

def debug_whitelist():
    console = Console()
    console.print("[bold blue]Debugging Property Whitelist Matching...[/bold blue]\n")
    
    file_path = project_root / "data" / "CRES - Portfolio Database.xlsm"
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    # Read whitelist from DB sheet
    console.print("[bold]1. Reading Whitelist from DB sheet (Column K):[/bold]")
    
    whitelist = []
    if 'DB' in wb.sheetnames:
        db_sheet = wb['DB']
        row_idx = 2
        while row_idx < 100:
            cell_val = db_sheet.cell(row=row_idx, column=11).value
            if not cell_val or (isinstance(cell_val, str) and not cell_val.strip()):
                break
            whitelist.append(str(cell_val).strip())
            row_idx += 1
    
    console.print(f"Found {len(whitelist)} properties in whitelist")
    
    # Show first 10
    wl_table = Table(title="Whitelist (first 10)")
    wl_table.add_column("Index")
    wl_table.add_column("Property Name")
    wl_table.add_column("Length")
    wl_table.add_column("Repr")
    
    for i, prop in enumerate(whitelist[:10]):
        wl_table.add_row(str(i), prop, str(len(prop)), repr(prop))
    console.print(wl_table)
    
    # Find all property pairs
    console.print("\n[bold]2. Finding Property Pairs:[/bold]")
    all_sheets = wb.sheetnames
    pairs = []
    for s in all_sheets:
        if s.endswith("-Fin"):
            base_name = s[:-4]
            bgt_sheet = f"{base_name}-Bgt"
            if bgt_sheet in all_sheets:
                pairs.append((base_name, s, bgt_sheet))
    
    console.print(f"Found {len(pairs)} property pairs")
    
    # Check specific properties
    console.print("\n[bold]3. Checking Specific Properties:[/bold]")
    target_props = ["Flats on maple", "Pecan Creek"]
    
    for target in target_props:
        console.print(f"\n[cyan]Checking: {repr(target)}[/cyan]")
        
        # Check if in whitelist
        in_whitelist = target in whitelist
        console.print(f"  In whitelist: {in_whitelist}")
        
        # Check if sheets exist
        fin_sheet = f"{target}-Fin"
        bgt_sheet = f"{target}-Bgt"
        has_sheets = fin_sheet in all_sheets and bgt_sheet in all_sheets
        console.print(f"  Has sheets: {has_sheets}")
        
        # Check exact match in pairs
        in_pairs = any(base == target for base, _, _ in pairs)
        console.print(f"  In pairs: {in_pairs}")
        
        # Try to find close matches in whitelist
        console.print("  Close matches in whitelist:")
        for wl_prop in whitelist:
            if target.lower() in wl_prop.lower() or wl_prop.lower() in target.lower():
                console.print(f"    - {repr(wl_prop)} (exact={wl_prop == target})")
    
    wb.close()

if __name__ == "__main__":
    debug_whitelist()
