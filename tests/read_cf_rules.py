"""Read conditional formatting rules from CRES Portfolio Internal sheet."""
import openpyxl
from pathlib import Path

def read_conditional_formatting():
    file_path = Path(r"c:\Users\VINCY\OneDrive - Vincy R Dsilva\VBA\Brendan\AI Project\OpenAI\data\CRES - Portfolio Database.xlsm")
    
    print(f"Loading workbook: {file_path.name}")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    # Find the CRES - Portfolio (Internal) sheet
    target_sheet = None
    for name in wb.sheetnames:
        if "Portfolio" in name and "Internal" in name:
            target_sheet = name
            break
    
    if not target_sheet:
        print("Could not find 'CRES - Portfolio (Internal)' sheet.")
        print(f"Available sheets: {wb.sheetnames[:10]}...")
        return
    
    print(f"\nFound sheet: {target_sheet}")
    ws = wb[target_sheet]
    
    # Read header row (Row 4, Columns B to AD)
    print("\n=== HEADERS (Row 4, B to AD) ===")
    headers = {}
    for col_idx in range(2, 31):  # B=2 to AD=30
        cell = ws.cell(row=4, column=col_idx)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            headers[col_letter] = cell.value
            print(f"  {col_letter}: {cell.value}")
    
    # Read conditional formatting rules
    print("\n=== CONDITIONAL FORMATTING RULES ===")
    cf = ws.conditional_formatting
    
    if not cf:
        print("No conditional formatting found.")
    else:
        for cf_range in cf:
            print(f"\nRange: {cf_range}")
            for rule in cf[cf_range]:
                print(f"  Rule Type: {rule.type}")
                if hasattr(rule, 'formula') and rule.formula:
                    print(f"  Formula: {rule.formula}")
                if hasattr(rule, 'colorScale') and rule.colorScale:
                    cs = rule.colorScale
                    print(f"  Color Scale: {cs}")
                if hasattr(rule, 'dataBar') and rule.dataBar:
                    print(f"  Data Bar: {rule.dataBar}")
                if hasattr(rule, 'iconSet') and rule.iconSet:
                    print(f"  Icon Set: {rule.iconSet}")
                if hasattr(rule, 'dxf') and rule.dxf:
                    dxf = rule.dxf
                    if dxf.fill:
                        print(f"  Fill: {dxf.fill}")
                    if dxf.font:
                        print(f"  Font: {dxf.font}")
                if hasattr(rule, 'operator'):
                    print(f"  Operator: {rule.operator}")
                if hasattr(rule, 'text'):
                    print(f"  Text: {rule.text}")
    
    wb.close()

if __name__ == "__main__":
    read_conditional_formatting()
